import openpyxl
import sys

def ayuda():
    print("Para utilizar este programa deberas escribir los datos separados por espacios por la linea de comandos de la siguiente manera:")
    print("Uso: py examen.py [listar|crear|editar|eliminar] [código] [marca] [modelo] [precio] [kilometraje]")

def crear_vehiculo(codigo, marca, modelo, precio, kilometraje):
    try:
        libro = openpyxl.load_workbook('vehiculos.xlsx')

        listado_sheet = libro['listado']

        for row in listado_sheet.iter_rows(min_row=2, max_col=1, values_only=True):
            if codigo == row[0]:
                print(f"\nError: El vehiculo con código {codigo} ya existe en el listado.")
                return

        print(f"Código: {codigo}, Marca: {marca}, Modelo: {modelo}, Precio: {precio}, Kilometraje: {kilometraje}")
        nueva_fila = [codigo, marca, modelo, precio, kilometraje]
        listado_sheet.append(nueva_fila)

        # Guarda los cambios en el libro
        libro.save('vehiculos.xlsx')
        print(f"\n El Vehículo con código {codigo} se creó exitosamente en el listado.")

    except FileNotFoundError:
        print("El archivo 'vehiculos.xlsx' no existe. Crea el listado antes de agregar vehículos.")

    finally:
        if 'libro' in locals():
            libro.close()

def editar_vehículo(codigo, marca, modelo, precio, kilometraje):
    try:
        
        libro = openpyxl.load_workbook('vehiculos.xlsx')

       
        listado_sheet = libro['listado']

       
        for row in listado_sheet.iter_rows(min_row=2, max_col=1, values_only=True):
            if codigo == row[0]:
               
                updated_row = [codigo, marca, modelo, precio, kilometraje]

               
                idx = row[0]
                for i, r in enumerate(listado_sheet.iter_rows(min_row=2, max_col=1, values_only=True), start=2):
                    if r[0] == idx:
                        for j, val in enumerate(updated_row, start=1):
                            listado_sheet.cell(row=i, column=j, value=val)
                        break

                libro.save('vehiculos.xlsx')
                print(f"\nEl Vehículo con código {codigo} actualizado en el listado.")
                return

        print(f"\nError: No se encontró un vehículo con código {codigo} en el listado.")

    except FileNotFoundError:
        print("El archivo 'vehículo.xlsx' no existe. Crea el listado antes de editar el vehículo.")

    finally:
        
        if 'libro' in locals():
            libro.close()

def eliminar_vehículo(codigo):
    try:
       
        libro = openpyxl.load_workbook('vehiculos.xlsx')

       
        listado_sheet = libro['listado']

       
        for i, row in enumerate(listado_sheet.iter_rows(min_row=2, max_col=1, values_only=True), start=2):
            if codigo == row[0]:
                # Elimina la fila del producto
                listado_sheet.delete_rows(i, 1)

               
                libro.save('vehiculos.xlsx')
                print(f"\nVehículo con código {codigo} eliminado del listado.")
                return

        print(f"\nError: No se encontró un vehículo con código {codigo} en el listado.")

    except FileNotFoundError:
        print("El archivo 'vehiculos.xlsx' no existe. Crea el listado antes de eliminar el vehículo.")

    finally:
        if 'libro' in locals():
            libro.close()

def listar_vehículos():
    try:
    
        libro = openpyxl.load_workbook('vehiculos.xlsx')

        listado_sheet = libro['listado']

        print("\nLIstado de vehículos:")
        for row in listado_sheet.iter_rows(min_row=2, values_only=True):
            codigo, marca, modelo, precio, kilometraje = row
            print(f"Código: {codigo}, Marca: {marca}, Modelo: {modelo}, Precio: {precio}, Kilometraje: {kilometraje}")

    except FileNotFoundError:
        print("El archivo 'vehiculos.xlsx' no existe. Crea el listado antes de listar.")

    finally:
        if 'libro' in locals():
            libro.close()

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: py examen.py [listar|crear|editar|eliminar] [código] [marca] [modelo] [precio] [kilometraje]")
        sys.exit(1)

    opcion = sys.argv[1]

    if opcion == 'crear' and len(sys.argv) == 7:
        _, opcion, codigo, marca, modelo, precio, kilometraje = sys.argv
        crear_vehiculo(codigo, marca, modelo, float(precio), int(kilometraje))
    elif opcion == 'editar' and len(sys.argv) == 7:
        _, opcion, codigo, marca, modelo, precio, kilometraje = sys.argv
        editar_vehículo(codigo, marca, modelo, float(precio), int(kilometraje))
    elif opcion == 'eliminar' and len(sys.argv) == 3:
        _, opcion, codigo = sys.argv
        eliminar_vehículo(codigo)   
    elif opcion == 'listar':
        listar_vehículos()
    elif opcion == 'ayuda' or 'help':
        ayuda()
    else:
        print("Ocurrió un error intentelo de nuevo")

















